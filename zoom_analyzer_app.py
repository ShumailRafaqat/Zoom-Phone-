#!/usr/bin/env python3
"""
Zoom Phone Agent Performance Analyzer 
----------------------------------------------
- Inbound + Outbound both
- All times in EST (as Zoom log)
- Breaks converted from PKT → EST
- Unusual calling time = free time outside breaks (no call activity)
- Calls after 15:00 EST are ignored

Run:  streamlit run zoom_analyzer_app.py
pip install streamlit pandas openpyxl
"""

import io
from datetime import datetime, time, timedelta

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
KNOWN_AGENTS = {
    "Alina Amir",
    "Eva Roberts",
    "careflare911@gmail.com",
    "Mitchell Atwood",
    "Norah Linder",
    "Ayrlish Brown",
    "Gaby Alex",
    "Ayan Ali",
    "Jessi brown",
    "Ayesha Asif",     
    "Annas Khan",
}

# Breaks in EST (converted from PKT)
# PKT 21:30–21:50 → EST 12:30–12:50
# PKT 00:00–00:30 → EST 15:00–15:30
# Working PKT 19:00–03:00 → EST 10:00–18:00
BREAK1_START, BREAK1_END = time(12, 30), time(12, 50)
BREAK2_START, BREAK2_END = time(15, 0), time(15, 30)

# Cut-off: calls starting at or after 15:00 EST are ignored
CUTOFF_TIME = time(15, 0)

# ============================================================
# HELPERS
# ============================================================
def parse_duration(d):
    if pd.isna(d) or str(d).strip() in ["--", " --", ""]:
        return 0
    try:
        parts = str(d).strip().split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        return 0
    except Exception:
        return 0


def sec_to_hms(sec):
    try:
        if sec is None or (isinstance(sec, float) and np.isnan(sec)):
            sec = 0
        sec = int(float(sec))
    except Exception:
        sec = 0
    if sec <= 0:
        return "00:00:00"
    return f"{sec // 3600:02d}:{(sec % 3600) // 60:02d}:{sec % 60:02d}"


def overlap_with_breaks(start, end):
    """Seconds of [start, end] inside official EST break windows."""
    if pd.isna(start) or pd.isna(end) or end <= start:
        return 0.0
    total = 0.0
    for day_offset in (0, 1):
        day = (start + timedelta(days=day_offset)).date()
        for bstart, bend in ((BREAK1_START, BREAK1_END), (BREAK2_START, BREAK2_END)):
            bs = datetime.combine(day, bstart)
            be = datetime.combine(day, bend)
            ov_s = max(start, bs)
            ov_e = min(end, be)
            if ov_e > ov_s:
                total += (ov_e - ov_s).total_seconds()
    return total


def resolve_agent(row):
    """Attribute call to an agent for both outbound and inbound."""
    direction = str(row.get("Direction", ""))
    if "Outbound" in direction:
        name = row.get("From Name")
        if pd.notna(name) and str(name) in KNOWN_AGENTS:
            return str(name)
    if "Inbound" in direction:
        for col in ("To Name", "To Email", "Operator Name"):
            val = row.get(col)
            if pd.notna(val) and str(val) in KNOWN_AGENTS:
                return str(val)
    return None


def load_csv(uploaded_file):
    df = pd.read_csv(uploaded_file, encoding="utf-8-sig")
    if "NO." in df.columns:
        df = df[df["NO."].notna() & (df["NO."] != "")].copy()

    required = ["Direction", "Start Time", "Call Result", "Duration"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"CSV missing columns: {missing}")

    df["Start Time"] = pd.to_datetime(df["Start Time"])
    df["Duration_sec"] = df["Duration"].apply(parse_duration).astype(int)
    if "Wait Time" in df.columns:
        df["Wait_sec"] = df["Wait Time"].apply(parse_duration).fillna(0).astype(int)
    else:
        df["Wait_sec"] = 0

    # ========== ONLY CHANGE: ignore calls starting at/after 15:00 EST ==========
    df = df[df["Start Time"].dt.time < CUTOFF_TIME].copy()
    # ==========================================================================

    df["Agent"] = df.apply(resolve_agent, axis=1)
    df = df[df["Agent"].notna()].copy()

    df["is_connected"] = df["Call Result"].isin(["Connected", "Answered"])
    df["is_failed"] = df["Call Result"].isin(
        ["Call Failed", "Canceled", "No Answer", "Ring Timeout", "Rejected", "Abandoned"]
    )
    df["talk_sec"] = np.where(df["is_connected"], df["Duration_sec"], 0).astype(int)
    df["end_time"] = df["Start Time"] + pd.to_timedelta(df["Duration_sec"], unit="s")
    df["date"] = df["Start Time"].dt.date
    df["Direction_clean"] = df["Direction"].astype(str).str.replace(r"\s*\(Internal\)", "", regex=True)

    def other_phone(row):
        d = str(row.get("Direction", ""))
        if "Outbound" in d:
            return str(row["To Phone Number"]) if pd.notna(row.get("To Phone Number")) else ""
        if "Inbound" in d:
            return str(row["From Phone Number"]) if pd.notna(row.get("From Phone Number")) else ""
        return ""

    def other_name(row):
        d = str(row.get("Direction", ""))
        if "Outbound" in d:
            return str(row["To Name"]) if pd.notna(row.get("To Name")) else ""
        if "Inbound" in d:
            return str(row["From Name"]) if pd.notna(row.get("From Name")) else ""
        return ""

    df["Other Phone"] = df.apply(other_phone, axis=1)
    df["Other Name"] = df.apply(other_name, axis=1)
    df = df.sort_values(["Agent", "Start Time"]).reset_index(drop=True)
    return df


def compute_unusual_gaps(agent_df):
    """Unusual free time BEFORE each call = gap minus break overlap.
    Returns list of unusual seconds before each call + total.
    """
    agent_df = agent_df.sort_values("Start Time").copy().reset_index(drop=True)
    unusual_before = []
    total_unusual = 0.0
    for i in range(len(agent_df)):
        if i == 0:
            unusual_before.append(0.0)
            continue
        prev_end = agent_df.iloc[i - 1]["end_time"]
        this_start = agent_df.iloc[i]["Start Time"]
        gap = (this_start - prev_end).total_seconds()
        if pd.isna(gap) or gap <= 0:
            unusual_before.append(0.0)
            continue
        break_ov = overlap_with_breaks(prev_end, this_start)
        unusual = max(0.0, gap - break_ov)
        unusual_before.append(unusual)
        total_unusual += unusual
    return unusual_before, total_unusual


def analyze_unusual_gaps(agent_df):
    """Detailed unusual gap analysis for Agent Summary.
    Returns total, counts of significant gaps, and longest gap info.
    """
    agent_df = agent_df.sort_values("Start Time").copy().reset_index(drop=True)
    total_unusual = 0.0
    gaps_gt_3 = 0
    gaps_gt_5 = 0
    gaps_gt_10 = 0
    longest = 0.0
    longest_timing = "—"

    for i in range(1, len(agent_df)):
        prev_end = agent_df.iloc[i - 1]["end_time"]
        this_start = agent_df.iloc[i]["Start Time"]
        gap = (this_start - prev_end).total_seconds()
        if pd.isna(gap) or gap <= 0:
            continue
        break_ov = overlap_with_breaks(prev_end, this_start)
        unusual = max(0.0, gap - break_ov)
        total_unusual += unusual

        if unusual > 180:
            gaps_gt_3 += 1
        if unusual > 300:
            gaps_gt_5 += 1
        if unusual > 600:
            gaps_gt_10 += 1

        if unusual > longest:
            longest = unusual
            # Full timing: previous call end – next call start
            longest_timing = f"{prev_end.strftime('%H:%M')} – {this_start.strftime('%H:%M')}"

    return {
        "total_unusual": total_unusual,
        "gaps_gt_3": gaps_gt_3,
        "gaps_gt_5": gaps_gt_5,
        "gaps_gt_10": gaps_gt_10,
        "longest": longest,
        "longest_timing": longest_timing,
    }


def build_all_calls(df):
    rows = []
    for agent in sorted(KNOWN_AGENTS):
        ad = df[df["Agent"] == agent].sort_values("Start Time").copy().reset_index(drop=True)
        if len(ad) == 0:
            continue
        ad["_phone"] = ad["Other Phone"].fillna("UNKNOWN").astype(str)
        ad["attempt_on_number"] = ad.groupby("_phone").cumcount() + 1
        unusual_list, _ = compute_unusual_gaps(ad)

        for i in range(len(ad)):
            row = ad.iloc[i]
            unusual = unusual_list[i]
            raw_gap = 0
            if i > 0:
                raw_gap = (row["Start Time"] - ad.iloc[i - 1]["end_time"]).total_seconds()
                if pd.isna(raw_gap) or raw_gap < 0:
                    raw_gap = 0
            rows.append(
                {
                    "Agent": agent,
                    "Date": str(row["Start Time"].date()),
                    "Direction": row["Direction_clean"],
                    "Call Result": str(row["Call Result"]) if pd.notna(row["Call Result"]) else "",
                    "Other Phone": row["Other Phone"],
                    "Other Name": row["Other Name"],
                    "Start Time": row["Start Time"].strftime("%Y-%m-%d %H:%M:%S"),
                    "End Time": row["end_time"].strftime("%Y-%m-%d %H:%M:%S"),
                    "Duration": sec_to_hms(row["Duration_sec"]),
                    "Duration (sec)": int(row["Duration_sec"]),
                    "Wait Time": sec_to_hms(row["Wait_sec"]),
                    "Talk (sec)": int(row["talk_sec"]),
                    "Gap from Previous (sec)": int(raw_gap),
                    "Unusual Calling Time (sec)": int(unusual),
                    "Unusual Calling Time": sec_to_hms(unusual) if unusual > 0 else "—",
                    "Attempt # on this Number": int(row["attempt_on_number"]),
                    "Call ID": str(row["Call ID"]) if "Call ID" in ad.columns and pd.notna(row.get("Call ID")) else "",
                }
            )
    return pd.DataFrame(rows)


def build_daily_summary(df):
    rows = []
    for day in sorted(df["date"].unique()):
        day_df = df[df["date"] == day]
        for agent in sorted(KNOWN_AGENTS):
            ad = day_df[day_df["Agent"] == agent].sort_values("Start Time")
            if len(ad) == 0:
                continue
            total_calls = len(ad)
            outbound = int(ad["Direction_clean"].str.contains("Outbound", na=False).sum())
            inbound = int(ad["Direction_clean"].str.contains("Inbound", na=False).sum())
            connected = int(ad["is_connected"].sum())
            failed = int(ad["is_failed"].sum())
            unique_nums = int(ad["Other Phone"].replace("", np.nan).nunique())
            attended_nums = int(
                ad.loc[ad["is_connected"], "Other Phone"].replace("", np.nan).nunique()
            )
            total_talk = int(ad["talk_sec"].sum())
            multi_nums = int((ad.groupby("Other Phone").size() > 1).sum()) if len(ad) else 0

            # Detailed unusual gap analysis
            analysis = analyze_unusual_gaps(ad)

            rows.append(
                {
                    "Agent": agent,
                    "Date": str(day),
                    "Total Calls": total_calls,
                    "Outbound": outbound,
                    "Inbound": inbound,
                    "Connected": connected,
                    "Failed/Canceled": failed,
                    "Connect Rate %": round(100.0 * connected / total_calls, 1) if total_calls else 0.0,
                    "Unique Numbers": unique_nums,
                    "Numbers Attended (Connected)": attended_nums,
                    "Numbers with >1 Call": multi_nums,
                    "Total Talk Time": sec_to_hms(total_talk),
                    "Total Talk (sec)": total_talk,
                    "Unusual Calling Time": sec_to_hms(analysis["total_unusual"]),
                    "Unusual Calling Time (sec)": int(analysis["total_unusual"]),
                    "Gaps >3 min": analysis["gaps_gt_3"],
                    "Gaps >5 min": analysis["gaps_gt_5"],
                    "Gaps >10 min": analysis["gaps_gt_10"],
                    "Longest Gap": sec_to_hms(analysis["longest"]) if analysis["longest"] > 0 else "—",
                    "Longest Gap Timing": analysis["longest_timing"],
                    "First Call": ad["Start Time"].min().strftime("%H:%M:%S"),
                    "Last Call": ad["Start Time"].max().strftime("%H:%M:%S"),
                }
            )
    return pd.DataFrame(rows)


def build_multi(df):
    count_col = "Call ID" if "Call ID" in df.columns else "Start Time"
    g = (
        df[df["Other Phone"].astype(str).str.len() > 0]
        .groupby(["Agent", "Other Phone"])
        .agg(
            Total_Attempts=(count_col, "count"),
            Connected=("is_connected", "sum"),
            Failed=("is_failed", "sum"),
            Outbound=("Direction_clean", lambda x: int(x.str.contains("Outbound").sum())),
            Inbound=("Direction_clean", lambda x: int(x.str.contains("Inbound").sum())),
            Total_Talk_sec=("talk_sec", "sum"),
            First_Call=("Start Time", "min"),
            Last_Call=("Start Time", "max"),
            Results=("Call Result", lambda x: " | ".join(x.astype(str).value_counts().head(4).index.tolist())),
        )
        .reset_index()
    )
    g = g[g["Total_Attempts"] >= 2].copy()
    g["Connect Rate %"] = (100 * g["Connected"] / g["Total_Attempts"]).round(1)
    g["Total Talk"] = g["Total_Talk_sec"].apply(sec_to_hms)
    g["Span Hours"] = ((g["Last_Call"] - g["First_Call"]).dt.total_seconds() / 3600).round(2)
    g["First Call"] = g["First_Call"].dt.strftime("%Y-%m-%d %H:%M:%S")
    g["Last Call"] = g["Last_Call"].dt.strftime("%Y-%m-%d %H:%M:%S")
    return g.sort_values(["Agent", "Total_Attempts"], ascending=[True, False])


# ============================================================
# EXCEL
# ============================================================
def style_header(ws, row, sc, ec):
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(sc, ec + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin


def auto_width(ws, min_w=8, max_w=22):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row + 1, 80)):
            try:
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 1, min_w), max_w)


def write_report(daily_df, multi, all_calls_df) -> bytes:
    wb = Workbook()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt = PatternFill("solid", fgColor="E8F4FD")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green = PatternFill("solid", fgColor="C6EFCE")
    orange = PatternFill("solid", fgColor="FCE4D6")

    # Unusual Calling Time colors (All Calls Detail only)
    yellow_light = PatternFill("solid", fgColor="FFF2CC")   # < 3 min
    soft_pink    = PatternFill("solid", fgColor="E8A0A8")   # > 3 min (thora dark)
    magenta      = PatternFill("solid", fgColor="FF00FF")   # > 5 min
    bright_red   = PatternFill("solid", fgColor="FF0000")   # > 10 min

    title_font = Font(bold=True, size=13, color="1F4E79")
    section_font = Font(bold=True, size=11, color="2F5496")

    # ========== 1. Agent Summary ==========
    ws1 = wb.active
    ws1.title = "Agent Summary"
    ws1["A1"] = "Agent Summary (Inbound + Outbound) — Times in EST | Calls after 15:00 ignored"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:U1")
    ws1["A2"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        "Log: EST/EDT | Working EST 10:00–15:00 (cutoff) | Breaks EST 12:30–12:50 & 15:00–15:30 "
        "(= PKT 19:00–03:00 / 21:30–21:50 / 00:00–00:30)"
    )
    ws1["A2"].font = Font(italic=True, size=9)
    ws1.merge_cells("A2:U2")
    ws1["A3"] = (
        "Unusual Calling Time = free time with NO call activity, AFTER removing official breaks. "
        "Gaps >3/5/10 min = how many times agent was free that long. "
        "Longest Gap Timing = previous call end – next call start (when biggest free gap happened). "
        "NOTE: Any call starting at/after 15:00 EST is completely ignored."
    )
    ws1["A3"].font = Font(size=9, color="666666")
    ws1.merge_cells("A3:U3")

    sum_cols = [
        "Agent", "Date", "Total Calls", "Outbound", "Inbound", "Connected", "Failed/Canceled",
        "Connect Rate %", "Unique Numbers", "Numbers Attended (Connected)", "Numbers with >1 Call",
        "Total Talk Time", "Unusual Calling Time",
        "Gaps >3 min", "Gaps >5 min", "Gaps >10 min",
        "Longest Gap", "Longest Gap Timing",
        "First Call", "Last Call",
    ]
    for c, h in enumerate(sum_cols, 1):
        ws1.cell(row=5, column=c, value=h)
    style_header(ws1, 5, 1, len(sum_cols))
    ws1.row_dimensions[5].height = 32

    for r_idx, (_, row) in enumerate(daily_df.iterrows(), 6):
        vals = [row[c] for c in sum_cols]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r_idx, column=c, value=v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if r_idx % 2 == 0:
                cell.fill = alt
            # Highlight high unusual time
            if sum_cols[c - 1] == "Unusual Calling Time" and row["Unusual Calling Time (sec)"] > 1800:
                cell.fill = red
            elif sum_cols[c - 1] == "Unusual Calling Time" and row["Unusual Calling Time (sec)"] > 600:
                cell.fill = yellow
            # Highlight high gap counts
            if sum_cols[c - 1] == "Gaps >10 min" and row["Gaps >10 min"] >= 1:
                cell.fill = bright_red
            elif sum_cols[c - 1] == "Gaps >5 min" and row["Gaps >5 min"] >= 2:
                cell.fill = magenta
            elif sum_cols[c - 1] == "Gaps >3 min" and row["Gaps >3 min"] >= 3:
                cell.fill = soft_pink
    auto_width(ws1)
    ws1.freeze_panes = "C6"

    # ========== 2. All Calls Detail ==========
    ws2 = wb.create_sheet("All Calls Detail")
    ws2["A1"] = "All Calls Detail — Inbound + Outbound (EST) | Calls after 15:00 ignored"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:Q1")
    ws2["A2"] = (
        "Green=Connected, Red=Failed. "
        "Unusual Calling Time = free gap BEFORE this call (outside breaks). "
        "Colors: Yellow(<3min) | Soft Pink(>3min) | Magenta(>5min) | Bright Red(>10min)"
    )
    ws2["A2"].font = Font(italic=True, size=9)
    ws2.merge_cells("A2:Q2")

    call_cols = [
        "Agent", "Date", "Direction", "Call Result", "Other Phone", "Other Name",
        "Start Time", "End Time", "Duration", "Duration (sec)", "Wait Time", "Talk (sec)",
        "Gap from Previous (sec)", "Unusual Calling Time", "Unusual Calling Time (sec)",
        "Attempt # on this Number", "Call ID",
    ]
    call_cols = [c for c in call_cols if c in all_calls_df.columns]
    for c, h in enumerate(call_cols, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header(ws2, 4, 1, len(call_cols))
    ws2.row_dimensions[4].height = 30

    for r_idx in range(len(all_calls_df)):
        result = all_calls_df.iloc[r_idx]["Call Result"]
        unusual_sec = int(all_calls_df.iloc[r_idx].get("Unusual Calling Time (sec)", 0) or 0)
        for c_idx, col in enumerate(call_cols, 1):
            val = all_calls_df.iloc[r_idx][col]
            cell = ws2.cell(row=r_idx + 5, column=c_idx, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if result in ("Connected", "Answered"):
                cell.fill = green
            elif result in ("Call Failed", "Canceled", "No Answer", "Ring Timeout", "Rejected", "Abandoned"):
                cell.fill = red
            elif (r_idx + 5) % 2 == 0:
                cell.fill = alt
            # Unusual Calling Time highlight (only these two columns)
            if col in ("Unusual Calling Time", "Unusual Calling Time (sec)") and unusual_sec > 0:
                if unusual_sec > 600:          # > 10 minutes
                    cell.fill = bright_red
                elif unusual_sec > 300:        # > 5 minutes
                    cell.fill = magenta
                elif unusual_sec > 180:        # > 3 minutes
                    cell.fill = soft_pink
                else:                          # < 3 minutes (but > 0)
                    cell.fill = yellow_light
    auto_width(ws2, max_w=18)
    ws2.freeze_panes = "D5"

    # ========== 3. Multi-Attempt Numbers ==========
    ws3 = wb.create_sheet("Multi-Attempt Numbers")
    ws3["A1"] = "Multi-Attempt Numbers — Numbers contacted 2+ times by same agent"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:L1")
    ws3["A2"] = "Which numbers were contacted multiple times, first/last time, talk, connect rate. Yellow = 5+ attempts."
    ws3["A2"].font = Font(italic=True, size=9)
    ws3.merge_cells("A2:L2")

    mh = [
        "Agent", "Other Phone", "Total_Attempts", "Outbound", "Inbound",
        "Connected", "Failed", "Connect Rate %", "Total Talk",
        "First Call", "Last Call", "Span Hours", "Results",
    ]
    mh = [c for c in mh if c in multi.columns]
    labels = {
        "Total_Attempts": "Total Attempts",
        "Other Phone": "Phone Number",
        "Span Hours": "Span (hours)",
        "Results": "Results Mix",
        "Total Talk": "Total Talk Time",
    }
    for c, h in enumerate(mh, 1):
        ws3.cell(row=4, column=c, value=labels.get(h, h))
    style_header(ws3, 4, 1, len(mh))

    r = 5
    for _, row in multi.iterrows():
        vals = [row[c] for c in mh]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if r % 2 == 0:
                cell.fill = alt
            if mh[c - 1] == "Total_Attempts" and row["Total_Attempts"] >= 5:
                cell.fill = yellow
        r += 1
    auto_width(ws3)
    ws3.freeze_panes = "C5"

    # ========== 4. Definitions ==========
    ws4 = wb.create_sheet("Definitions")
    ws4["A1"] = "Definitions"
    ws4["A1"].font = title_font
    notes = [
        "",
        "TIMEZONE",
        "  • All call times shown in EST/EDT (same as Zoom Phone log).",
        "  • Working hours considered: EST 10:00–15:00 (calls after 15:00 are ignored)",
        "  • Break 1 (PKT 21:30–21:50) = EST 12:30–12:50",
        "  • Break 2 (PKT 00:00–00:30) = EST 15:00–15:30",
        "",
        "CUT-OFF RULE",
        "  • Any call whose Start Time is 15:00 EST or later is completely ignored.",
        "  • No analysis, no unusual time, no summary — simply removed.",
        "",
        "UNUSUAL CALLING TIME",
        "  • Free time when agent had NO call activity (no inbound, no outbound).",
        "  • Official break windows are EXCLUDED (not counted as unusual).",
        "  • Even short free gaps (seconds) outside breaks are counted in total.",
        "  • On All Calls Detail: Unusual Calling Time on a row = free gap BEFORE that call.",
        "  • On Agent Summary: total unusual free time for that agent that day.",
        "",
        "AGENT SUMMARY – NEW INSIGHT COLUMNS",
        "  • Gaps >3 min  = how many times agent was free more than 3 minutes",
        "  • Gaps >5 min  = how many times agent was free more than 5 minutes",
        "  • Gaps >10 min = how many times agent was free more than 10 minutes",
        "  • Longest Gap  = the single biggest free gap of the day",
        "  • Longest Gap Timing = previous call end time – next call start time (when biggest gap happened)",
        "  → One view se clear: agent ne kab aur kitna time waste kiya.",
        "",
        "OTHER METRICS",
        "  • Talk Time = Duration of Connected / Answered calls only.",
        "  • Failed/Canceled counted in volume but not as talk or unusual time.",
        "  • Numbers Attended = unique numbers with at least one Connected call.",
        "  • Attempt # = running count of contacts with that number by this agent.",
        "",
        "SHEETS",
        "  1. Agent Summary – daily totals + gap insights per agent",
        "  2. All Calls Detail – every inbound + outbound call with gaps & unusual time",
        "  3. Multi-Attempt Numbers – numbers contacted 2+ times",
        "  4. Definitions",
        "",
        "COLORS (All Calls Detail - Unusual Calling Time):",
        "  • Light Yellow (#FFF2CC) = less than 3 minutes",
        "  • Soft Pink (#E8A0A8)    = greater than 3 minutes",
        "  • Magenta (#FF00FF)      = greater than 5 minutes",
        "  • Bright Red (#FF0000)   = greater than 10 minutes",
        "  Green=Connected | Red=Failed | Yellow=5+ attempts (Multi sheet)",
    ]
    for i, line in enumerate(notes, 3):
        ws4.cell(row=i, column=1, value=line)
        if any(line.startswith(x) for x in ("TIMEZONE", "CUT-OFF", "UNUSUAL", "AGENT SUMMARY", "OTHER", "SHEETS", "COLORS")):
            ws4.cell(row=i, column=1).font = section_font
    ws4.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# STREAMLIT UI
# ============================================================
st.set_page_config(page_title="Zoom Agent Performance", page_icon="📞", layout="wide")

st.title("📞 Zoom Phone – Agent Performance Analyzer")
st.caption(
    "Inbound + Outbound · Calls after 15:00 EST ignored"
)

with st.expander("Working hours & breaks (EST)", expanded=False):
    st.markdown(
        """
        - **Zoom log timezone:** EST/EDT (all call times shown as-is)  
        - **Working hours considered:** EST 10:00–15:00 (**calls starting at/after 15:00 are ignored**)  
        - **Break 1:** PKT 21:30–21:50 → **EST 12:30–12:50**  
        - **Break 2:** PKT 00:00–00:30 → **EST 15:00–15:30**  
        - **Unusual Calling Time:** agent free (no inbound/outbound) outside breaks — highlighted  
        """
    )

uploaded = st.file_uploader("Upload Zoom Phone call log CSV", type=["csv"])

if uploaded is not None:
    with st.spinner("Processing inbound + outbound…"):
        try:
            df = load_csv(uploaded)
            if len(df) == 0:
                st.error("No calls found for known agents (inbound or outbound) before 15:00 EST.")
                st.stop()

            daily_df = build_daily_summary(df)
            multi = build_multi(df)
            all_calls_df = build_all_calls(df)
            excel_bytes = write_report(daily_df, multi, all_calls_df)

            st.success(
                f"Done! {len(daily_df)} agent-day rows · "
                f"{len(all_calls_df)} calls (in+out) · "
                f"{len(multi)} multi-attempt numbers"
            )

            st.subheader("Agent Summary preview")
            preview = [
                "Agent", "Date", "Total Calls", "Outbound", "Inbound", "Connected",
                "Unique Numbers", "Total Talk Time", "Unusual Calling Time",
                "Gaps >3 min", "Gaps >5 min", "Gaps >10 min",
                "Longest Gap", "Longest Gap Timing",
            ]
            st.dataframe(daily_df[preview], use_container_width=True)

            st.download_button(
                label="⬇️ Download Full Excel Report",
                data=excel_bytes,
                file_name=f"Agent_Performance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Error: {e}")
            st.exception(e)
else:
    st.info("👆 Upload a Zoom Phone call-log CSV to start.")
